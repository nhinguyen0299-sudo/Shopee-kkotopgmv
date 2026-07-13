from flask import Flask, request, send_file, jsonify
from flask_cors import CORS
import pandas as pd
import re, io, os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

app = Flask(__name__)
CORS(app, origins="*", allow_headers=["Content-Type"], methods=["GET", "POST", "OPTIONS"])

KAKAO = ['kakaonline2','kakaonline3','kakaonline4','kakaonline7','kakaonline8']
DARK="1F3864"; MID="2F5496"; LIGHT="D6E4F7"; WHITE="FFFFFF"; GRAY="F2F5FA"

def clean(v):
    return re.sub(r'[\x00-\x08\x0b\x0c\x0e-\x1f]', '', v) if isinstance(v, str) else v

def hdr(cell, bg=DARK, fg=WHITE, size=10):
    cell.font = Font(name='Arial', bold=True, color=fg, size=size)
    cell.fill = PatternFill('solid', start_color=bg)
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

def brd(cell):
    s = Side(style='thin', color="BBBBBB")
    cell.border = Border(left=s, right=s, top=s, bottom=s)

def build_excel(df_all, months):
    agg = df_all.groupby(['Tên tài khoản KOL','Tháng','Tên Shop','ID Shop','Tên sản phẩm','Mã sản phẩm','Danh mục sản phẩm cấp 1'])['Giá trị mua hàng(₫)'].sum().reset_index()
    comm = df_all.groupby(['Tên tài khoản KOL','Tháng','Tên Shop','ID Shop','Tên sản phẩm','Mã sản phẩm','Danh mục sản phẩm cấp 1'])['Tỷ lệ hoa hồng người bán'].agg(lambda x: x.mode().iloc[0] if len(x.mode())>0 else '').reset_index()
    agg = agg.merge(comm, on=['Tên tài khoản KOL','Tháng','Tên Shop','ID Shop','Tên sản phẩm','Mã sản phẩm','Danh mục sản phẩm cấp 1'])
    agg.columns = ['Kênh','Tháng','Tên Shop','ID Shop','Tên sản phẩm','Mã sản phẩm','Ngành hàng','Tổng GMV','Tỷ lệ HH']
    agg = agg.sort_values(['Kênh','Tháng','Tổng GMV'], ascending=[True,True,False])
    agg['Rank'] = agg.groupby(['Kênh','Tháng']).cumcount() + 1
    for c in ['Tên Shop','Tên sản phẩm','Ngành hàng']: agg[c] = agg[c].apply(clean)

    wb = Workbook()

    DASH_COLS = ['#','Tên Shop','ID Shop','Ngành hàng','Tên sản phẩm','Mã sản phẩm','Tổng GMV (₫)','Tỷ lệ HH NB','Link sản phẩm']
    channel_options = sorted(KAKAO) + ['Tất cả kênh']

    # ── 1 sheet per channel + "Tất cả kênh" ─────────────────────────────────
    # Build "all" aggregation
    agg_all = agg.groupby(['Tháng','Tên Shop','ID Shop','Ngành hàng','Tên sản phẩm','Mã sản phẩm'])['Tổng GMV'].sum().reset_index()
    comm_all = agg.groupby(['Tháng','Tên Shop','ID Shop','Ngành hàng','Tên sản phẩm','Mã sản phẩm'])['Tỷ lệ HH'].agg(lambda x: x.mode().iloc[0] if len(x.mode())>0 else '').reset_index()
    agg_all = agg_all.merge(comm_all, on=['Tháng','Tên Shop','ID Shop','Ngành hàng','Tên sản phẩm','Mã sản phẩm'])
    agg_all = agg_all.sort_values(['Tháng','Tổng GMV'], ascending=[True,False])

    def write_data_sheet(wb, sheet_name, data_df, tab_color):
        ws = wb.create_sheet(sheet_name)
        ws.sheet_properties.tabColor = tab_color
        ws.freeze_panes = 'A2'
        ws.row_dimensions[1].height = 28
        for c, h in enumerate(DASH_COLS, 1):
            hdr(ws.cell(1, c, h), bg="444444")
        for ri, row in enumerate(data_df.itertuples(index=False), 2):
            # Unpack depending on columns
            if len(row) == 8:  # per-channel: Tháng,Shop,ID,Ngành,SP,Mã,GMV,HH
                thang,shop,sid_,tensp,masp_,nganh,gmv,hh = row
            else:
                thang,shop,sid_,nganh,tensp,masp_,gmv,hh = row
            sid  = str(int(sid_))  if pd.notna(sid_)  else ''
            masp = str(int(masp_)) if pd.notna(masp_) else ''
            link = f'https://shopee.vn/product/{sid}/{masp}' if sid and masp else ''
            rank = ri - 1
            bg = "FFD700" if rank==1 else "E8E8E8" if rank==2 else "CDAA7D" if rank==3 else (LIGHT if rank%2==0 else WHITE)

            rc = ws.cell(ri,1,rank)
            rc.font=Font(name='Arial',size=9,bold=(rank<=3))
            rc.fill=PatternFill('solid',start_color=bg)
            rc.alignment=Alignment(horizontal='center',vertical='center')
            brd(rc)

            vals = [clean(str(shop)), sid, clean(str(nganh)), clean(str(tensp)), masp, float(gmv) if pd.notna(gmv) else 0, str(hh) if pd.notna(hh) else '', link]
            for ci, val in enumerate(vals, 2):
                cell = ws.cell(ri, ci, val)
                cell.font = Font(name='Arial', size=9)
                cell.fill = PatternFill('solid', start_color=bg)
                cell.alignment = Alignment(vertical='center', wrap_text=(ci==5))
                brd(cell)
                if ci==7: cell.number_format='#,##0'
                if ci==8: cell.alignment=Alignment(horizontal='center',vertical='center')
                if ci==9 and link:
                    cell.hyperlink=link
                    cell.font=Font(name='Arial',size=9,color="0563C1",underline='single')

        for col,w in zip('ABCDEFGHI',[5,30,14,20,55,16,18,12,40]):
            ws.column_dimensions[col].width=w
        ws.row_dimensions[1].height=28

    # Write per-channel sheets (top 500 per month, use latest month)
    latest_month = months[-1]
    TAB = {'kakaonline2':'E74C3C','kakaonline3':'E67E22','kakaonline4':'27AE60','kakaonline7':'2980B9','kakaonline8':'8E44AD'}

    for ch in sorted(KAKAO):
        ch_data = agg[(agg['Kênh']==ch)&(agg['Tháng']==latest_month)].head(500)
        ch_data = ch_data[['Tháng','Tên Shop','ID Shop','Ngành hàng','Tên sản phẩm','Mã sản phẩm','Tổng GMV','Tỷ lệ HH']]
        write_data_sheet(wb, ch, ch_data, TAB.get(ch,'999999'))

    # Write "Tất cả kênh" sheet
    all_data = agg_all[agg_all['Tháng']==latest_month].head(500)
    write_data_sheet(wb, 'Tat ca kenh', all_data, '1F3864')

    # ── Dashboard ────────────────────────────────────────────────────────────
    ws_d = wb.create_sheet("Dashboard", 0)
    ws_d.sheet_properties.tabColor = DARK
    ws_d.freeze_panes = 'A8'

    ws_d.row_dimensions[1].height = 40
    ws_d.merge_cells('A1:I1')
    ws_d['A1'] = "TOP 500 GMV THEO KÊNH & THÁNG — KakaOnline"
    ws_d['A1'].font = Font(name='Arial', bold=True, size=16, color=WHITE)
    ws_d['A1'].fill = PatternFill('solid', start_color=DARK)
    ws_d['A1'].alignment = Alignment(horizontal='center', vertical='center')

    ws_d.row_dimensions[2].height = 22
    ws_d.merge_cells('A2:I2')
    ws_d['A2'] = f"Dữ liệu tháng {latest_month} — Upload CSV mới lên web app để cập nhật"
    ws_d['A2'].font = Font(name='Arial', italic=True, size=9, color="DDDDDD")
    ws_d['A2'].fill = PatternFill('solid', start_color=DARK)
    ws_d['A2'].alignment = Alignment(horizontal='center', vertical='center')

    ws_d.row_dimensions[3].height = 10

    # Channel tabs guide
    ws_d.row_dimensions[4].height = 28
    ws_d.merge_cells('A4:I4')
    ws_d['A4'] = "👇 Chọn sheet theo kênh bên dưới để xem Top 500 GMV từng kênh"
    ws_d['A4'].font = Font(name='Arial', bold=True, size=11, color=DARK)
    ws_d['A4'].alignment = Alignment(horizontal='center', vertical='center')

    # Channel quick nav boxes
    ws_d.row_dimensions[5].height = 40
    nav_channels = sorted(KAKAO) + ['Tat ca kenh']
    nav_labels   = sorted(KAKAO) + ['Tất cả kênh']
    nav_colors   = [TAB.get(c,'1F3864') for c in sorted(KAKAO)] + ['1F3864']
    for i, (ch, lbl, col) in enumerate(zip(nav_channels, nav_labels, nav_colors)):
        cell = ws_d.cell(5, i+1, lbl)
        cell.font = Font(name='Arial', bold=True, size=9, color=WHITE)
        cell.fill = PatternFill('solid', start_color=col)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        # Hyperlink to sheet
        cell.hyperlink = f"#{ch}!A1"

    ws_d.row_dimensions[6].height = 10
    ws_d.row_dimensions[7].height = 28

    # Summary table header
    ws_d.merge_cells('A7:I7')
    ws_d['A7'] = f"📊 Tổng hợp GMV theo kênh — Tháng {latest_month}"
    ws_d['A7'].font = Font(name='Arial', bold=True, size=12, color=WHITE)
    ws_d['A7'].fill = PatternFill('solid', start_color=MID)
    ws_d['A7'].alignment = Alignment(horizontal='center', vertical='center')

    # Summary data
    summary_headers = ['Kênh', 'Số sản phẩm', 'Tổng GMV (₫)', 'GMV Top 1', 'Sản phẩm Top 1']
    for c, h in enumerate(summary_headers, 1):
        cell = ws_d.cell(8, c, h)
        hdr(cell, bg=MID)
        brd(cell)

    row_idx = 9
    for ch in sorted(KAKAO):
        ch_data = agg[(agg['Kênh']==ch)&(agg['Tháng']==latest_month)]
        total_gmv = ch_data['Tổng GMV'].sum()
        n_products = len(ch_data)
        top1_gmv = ch_data['Tổng GMV'].iloc[0] if len(ch_data) > 0 else 0
        top1_name = clean(str(ch_data['Tên sản phẩm'].iloc[0])) if len(ch_data) > 0 else ''

        bg = LIGHT if row_idx % 2 == 0 else WHITE
        vals = [ch, n_products, total_gmv, top1_gmv, top1_name]
        for c, v in enumerate(vals, 1):
            cell = ws_d.cell(row_idx, c, v)
            cell.font = Font(name='Arial', size=10)
            cell.fill = PatternFill('solid', start_color=bg)
            cell.alignment = Alignment(vertical='center')
            brd(cell)
            if c in [3,4]: cell.number_format = '#,##0'
        row_idx += 1

    # Total row
    all_gmv = agg[agg['Tháng']==latest_month]['Tổng GMV'].sum()
    all_n   = len(agg_all[agg_all['Tháng']==latest_month])
    for c, v in enumerate(['TỔNG TẤT CẢ', all_n, all_gmv, '', ''], 1):
        cell = ws_d.cell(row_idx, c, v)
        cell.font = Font(name='Arial', bold=True, size=10, color=WHITE)
        cell.fill = PatternFill('solid', start_color=DARK)
        cell.alignment = Alignment(vertical='center')
        brd(cell)
        if c == 3: cell.number_format = '#,##0'

    for col,w in zip('ABCDE',[18,14,18,18,45]):
        ws_d.column_dimensions[col].width=w

    # Xóa sheet trống
    for name in list(wb.sheetnames):
        ws = wb[name]
        if ws.max_row<=1 and ws.max_column<=1:
            wb.remove(ws)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output

@app.route('/health', methods=['GET'])
def health():
    return jsonify({'status': 'ok'})

@app.route('/process', methods=['POST'])
def process():
    try:
        files = request.files.getlist('files')
        if not files:
            return jsonify({'error': 'Không có file nào được upload'}), 400

        dfs = []
        for f in files:
            try:
                df = pd.read_csv(f, low_memory=False, encoding='utf-8-sig')
                if df.shape[1] < 5: continue
                if 'Tên tài khoản KOL' not in df.columns: continue
                dfs.append(df)
            except Exception:
                continue

        if not dfs:
            return jsonify({'error': 'Không đọc được file CSV nào hợp lệ'}), 400

        df_all = pd.concat(dfs, ignore_index=True)
        df_all = df_all[df_all['Tên tài khoản KOL'].isin(KAKAO)].copy()

        if df_all.empty:
            return jsonify({'error': 'Không tìm thấy data của 5 kênh KakaOnline'}), 400

        df_all['Thời gian đặt hàng'] = pd.to_datetime(df_all['Thời gian đặt hàng'])
        df_all['Tháng'] = df_all['Thời gian đặt hàng'].dt.to_period('M').astype(str)

        for col in df_all.select_dtypes(include='object').columns:
            df_all[col] = df_all[col].apply(lambda x: clean(x) if isinstance(x,str) else x)

        months = sorted(df_all['Tháng'].unique().tolist())
        channels = df_all['Tên tài khoản KOL'].value_counts().to_dict()
        month = months[-1]  # latest month

        excel_file = build_excel(df_all, months)

        return send_file(
            excel_file,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'Top_GMV_KakaOnline_{month}.xlsx'
        )

    except Exception as e:
        return jsonify({'error': str(e)}), 500

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port)
